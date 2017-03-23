# -*- coding: utf-8 -*-
#! Obama Schedule 2012
import re
import requests
import json
import datetime, time
import openpyxl

from leader_list_getter_2016 import leader_name_list_getter

time1 = time.time()

#Column Setter
col = 'D'
# Year Setter
year = '2016'
#Pres Vp or Joint List
#[new_joint_meetings_list, new_PRES_MEETING_LIST, new_VP_MEETING_LIST]    
list_setter = 'new_joint_meetings_list'
#Wher my spreadsheet is opened and saved
excel_file_location = '/Users/cofax48/coding/QP_2016/Presidential_Visits.xlsx'

def web_to_file():
    """Download Leader's schedule and save as a text file"""
    urlL = []
    for i in range(1, 34): # 2014
        urlL.append('https://www.whitehouse.gov/blog?page=' + str(i))

    playFile = open('obama_schedule_{}.txt'.format(year), 'ab') 
    for q in urlL:
        res = requests.get(q) 
        res.raise_for_status()
        for chunk in res.iter_content(100000):
            playFile.write(chunk)            
    playFile.close()

'''
web_to_file()
'''
def country_name_acquistion():
    """A function that imports a list of countries and compiles it into a list"""
    filename = '/Users/cofax48/coding/QP_2016/countries_additional.json'
    with open(filename) as f:
        pop_data = json.load(f)
        
    country_name_list = ['G-20', 'G20', 'G8', 'G7', 'ASEAN', 'ASEAN-', 'APEC', 'G-8', 'NATO', 'CARICOM', 'Gulf Cooperation Council', 'GCC', 'G-7', 'Summit of the Americas', 'East Asia Summit', 'Strategic and Economic Dialogue']
    for pop_dict in pop_data:
        country_name = pop_dict["name"]
        country_name_list.append(country_name)
    country_dem_list = []
    filename = '/Users/cofax48/coding/QP_2016/indepth_country_list.json'
    with open(filename) as f:
        pop_data = json.load(f)
    for pop_dict in pop_data:
        country_demonym = pop_dict["name"]["common"], pop_dict["demonym"]
        country_name_list.append(country_demonym[1])
        country_dem_list.append(country_demonym)
    return {'country_name_list':country_name_list, 'country_dem_list':country_dem_list}

def daily_sched_search():
    """Searches the downloaded plaintxt file of Kerry's schedule"""
    open_sched_file = open('obama_schedule_{}.txt'.format(year))
    open_sched_file_r = open_sched_file.read()
    all_day_search = re.compile(r'\w{3,9} \d{1,2}, \d\d\d\d(?!,| is the|\.| on WhiteHouse| marks).*?•', re.DOTALL)
    ads = all_day_search.findall(open_sched_file_r)
    all_days = []

    #finds everyone's schedule for the whole day
    for days in ads:
        all_days.append(days)

    kerry_daily_meetings_list = []
    specific_day = re.compile(r'\w{3,9} \d{1,2}, \d\d\d\d')

    #finds Kerry's schedule within each particular day
    for i in all_days:
        sd_all_matches = specific_day.findall(i)
        sd = sd_all_matches[0]
        kerry_search = re.compile(r'\w{3,9} \d{1,2}, \d\d\d\d.*?•', re.DOTALL)
        kerry_daily_meeting = kerry_search.findall(i)
        for meets in kerry_daily_meeting:
            kerry_daily_meetings_list.append('\n\n' + str(sd) + '\n' + str(meets) )

    print('Searching the schedule')
    return {'kerry_daily_meetings_list':kerry_daily_meetings_list}

def meeting_search(kerry_daily_meetings_list, country_name_list):
    """Finds out who Kerry met with and from what country"""

    specific_day = re.compile(r'\w{3,9} \d{1,2}, \d\d\d\d')
    schedule_snippet = re.compile(r'(?!.*LOCAL TIME)\d{1,2}:\d{1,2} \w{2}.*?\n\n', re.DOTALL)
    old_bilateral_meetings_list = []

    meeting_name_list = ['met', 'mets', 'meet', 'meets', 'host', 'addresses', 'greets', 'welcomed', 'accompanies',
                         'greet', 'greeted', 'address', 'attend', 'attends', 'meeting', 'meetings', 'joins', 'has an audience with',
                         'joined', 'welcome', 'welcomes', 'welcomed', 'participates', 'deliver statements', 'delivers remarks',
                         'lunch', 'State Dinner']

    date_pattern = re.compile('\d{1,2}:\d{1,2} \w{2}.*?\d{1,2}:\d{1,2} \w{2}', re.DOTALL)
    for i in kerry_daily_meetings_list:
        sd_all_matches = specific_day.findall(i)
        sd = sd_all_matches[0]
        schedule_snippets = schedule_snippet.findall(i)   
        for snippet in schedule_snippets:
            for meeting in meeting_name_list:
                if str(meeting) in str(snippet):
                    whole_event = str(sd) + '\n' + str(snippet)
                    if whole_event in old_bilateral_meetings_list:
                        pass
                    else:
                        date_matches = date_pattern.findall(whole_event)
                        if len(date_matches) > 0:
                            specific_time = re.compile('\d{1,2}:\d{1,2} \w{2}')
                            each_time = specific_time.findall(whole_event)
                            new_date_string = whole_event.split(each_time[1])
                            sd_all_matches = specific_day.findall(whole_event)
                            sd = sd_all_matches[0]
                            new_whole_event = str(sd) + '\n' + str(new_date_string[1])
                            if str(new_date_string[0]) not in old_bilateral_meetings_list:
                                old_bilateral_meetings_list.append(new_date_string[0])
                            elif str(new_whole_event) not in old_bilateral_meetings_list:
                                old_bilateral_meetings_list.append(new_whole_event)
                        else:
                            old_bilateral_meetings_list.append(whole_event)


    bilateral_meetings_list = []
    for a in old_bilateral_meetings_list:
        if int(len(str(a))) < 400:
            if 'Northern Ireland' in str(a):
                a = a.replace('Northern Ireland', '')
                bilateral_meetings_list.append(a)
            elif 'Port of Spain' in str(a):
                a = a.replace('Port of Spain', '')
                bilateral_meetings_list.append(a)
            elif "Cote d’Ivoire" in str(a):
                a = a.replace("Cote d’Ivoire", 'Ivory Coast')
                bilateral_meetings_list.append(a)
            elif "Côte d'Ivoire" in str(a):
                a = a.replace("Côte d'Ivoire", 'Ivory Coast')
                bilateral_meetings_list.append(a)
            elif 'NATO Secretary-General Stoltenberg' in str(a):
                a = a.replace('NATO Secretary-General Stoltenberg', '')
                bilateral_meetings_list.append(a)                
            elif "Timor-Leste" in str(a):
                a = a.replace("Timor-Leste", 'East Timor')
                bilateral_meetings_list.append(a)
            elif "The Bahamas" in str(a):
                a = a.replace("The Bahamas", 'Bahamas, The')
                bilateral_meetings_list.append(a)
            elif 'Chairman and General Secretary of the National League of Democracy' in str(a):
                a = a.replace("Chairman and General Secretary of the National League of Democracy", 'Opposition Leader of Myanmar')
                bilateral_meetings_list.append(a)
            elif 'The Gambia' in str(a):
                a = a.replace("The Gambia", 'Gambia, The')
                bilateral_meetings_list.append(a)
            elif 'Panel for Sudan and South Sudan' in str(a):
                a = a.replace('Panel for Sudan and South Sudan', '')
                if 'former President' in str(a):
                    a = a.replace('former President', 'former Prez')
                    bilateral_meetings_list.append(a)
                else:
                    bilateral_meetings_list.append(a)
            elif "Sao Tome" in str(a):
                a = a.replace("Sao Tome and Principal", "São Tomé and Príncipal")
                bilateral_meetings_list.append(a)            
            elif 'Republic of the Congo' in str(a):
                a = a.replace('Republic of the Congo', 'Congo, Republic of the')
                bilateral_meetings_list.append(a)
            elif 'St.' in str(a):
                a = a.replace('St.', 'Saint')
                bilateral_meetings_list.append(a)
            elif 'will depart' in str(a):
                pass
            elif 'upcoming trip to ' in str(a):
                pass
            elif 'Ambassador to' in str(a):
                pass
            elif 'briefing to preview' in str(a):
                pass
            elif 'meets with his national security team' in str(a):
                pass
            elif 'hold a meeting on ' in str(a):
                pass
            elif 'former President' in str(a):
                a = a.replace('former President', 'former Prez')
                bilateral_meetings_list.append(a)
            elif 'G8' in str(a):
                a = a.replace('G8', 'G-8')
                bilateral_meetings_list.append(a)
            elif 'G20' in str(a):
                a = a.replace('G20', 'G-20')
                bilateral_meetings_list.append(a)
            elif 'Vice Premier' in str(a):
                a = a.replace('Vice Premier', 'Vice-Premeir')
                bilateral_meetings_list.append(a)
            elif 'former Prime Minister' in str(a):
                a = a.replace('former Prime Minister', 'former PM')
                bilateral_meetings_list.append(a)
            elif 'former British Prime Minister' in str(a):
                a = a.replace('former British Prime Minister', 'former PM of the United Kingdom')
                bilateral_meetings_list.append(a)
            elif 'ASEAN-' in str(a):
                a = a.replace('-', '')
                bilateral_meetings_list.append(a)
            elif 'Burma' in str(a):
                a = a.replace('Burma', 'Myanmar')
                bilateral_meetings_list.append(a)
            elif 'will remain' in str(a):
                pass
            elif 'ideo conference' in str(a):
                pass
            elif 'Deputy Prime Minister' in str(a):
                a = a.replace('Deputy Prime Minister', 'Deputy PM')
                bilateral_meetings_list.append(a)
            else:
                bilateral_meetings_list.append(a)


    country_meeting_list = []

    PRES_MEETINGS_LIST = []
    VP_MEETING_LIST = []
    joint_meetings_list = []
    for a in bilateral_meetings_list:
        if 'Vice President' in str(a):
            if 'Vice President will also attend' in str(a):
                joint_meetings_list.append(a)
            elif 'Vice President also attends' in str(a):
                joint_meetings_list.append(a)
            elif 'Vice President also attend' in str(a):
                joint_meetings_list.append(a)
            elif 'The President and Vice President' in str(a):
                joint_meetings_list.append(a)
            elif 'The President and the Vice President' in str(a):
                joint_meetings_list.append(a)                
            elif '; the Vice President' in str(a):
                joint_meetings_list.append(a)
            elif '; The Vice President' in str(a):
                joint_meetings_list.append(a)
            elif 'and the Vice President attends' in str(a):
                joint_meetings_list.append(a)                
            elif 'The President, the First Lady, and the Vice President' in str(a):
                joint_meetings_list.append(a)
            elif 'the President and the Vice Prez' in str(a):
                joint_meetings_list.append(a)
            elif 'The President, the Vice President' in str(a):
                joint_meetings_list.append(a)
            elif 'The President and the Vice President' in str(a):
                joint_meetings_list.append(a)
            elif 'The First Lady, the Vice Prez' in str(a):
                joint_meetings_list.append(a)
            else:
                VP_MEETING_LIST.append(a)
        else:
            PRES_MEETINGS_LIST.append(a)
            
    new_VP_MEETING_LIST = []
    for a in VP_MEETING_LIST:
        a = str(a).replace('Vice President', '', 1)
        a = str(a).replace('Vice President', 'Vice Prez')
        new_VP_MEETING_LIST.append(a)
    
    new_PRES_MEETING_LIST = []
    for a in PRES_MEETINGS_LIST:
        if 'he President' in str(a):
            a = a.replace('he President', '', 1)
            new_PRES_MEETING_LIST.append(a)
        elif 'President and the First Lady' in str(a):
            a = a.replace('President and the First Lady', '')
            new_PRES_MEETING_LIST.append(a)
        else:           
            new_PRES_MEETING_LIST.append(a)

    new_joint_meetings_list = []
    for a in joint_meetings_list:
        a = a.replace('The President', '')
        a = a.replace('the President', '')
        a = a.replace('the Vice President', '')
        a = a.replace('The President', '')
        new_joint_meetings_list.append(a)
        
        
    return {'new_joint_meetings_list':new_joint_meetings_list, 'new_PRES_MEETING_LIST':new_PRES_MEETING_LIST, 'new_VP_MEETING_LIST':new_VP_MEETING_LIST}

def country_and_leader_getter(new_joint_meetings_list, new_PRES_MEETING_LIST, new_VP_MEETING_LIST, country_name_list, country_dem_list):
    newest_country_and_list_thats_useable = leader_name_list_getter()['newest_country_and_list_thats_useable']
    new_country_leader_list = leader_name_list_getter()['new_country_leader_list']
    
    leader_list = ['President', 'Prime Minister', 'Prime Minster', 'King ', 'Chancellor', 'Taoiseach', 'Queen', 'Amir', 'Emperor', 'Pope', 
                   'CEO', 'Crown Prince', 'Defense Minister', 'Deputy Crown Prince', 'Deputy PM', 'Finance Minister', 'margins', 'leaders meeting', 'Working Session', 'Plenary Session', 'working session', 'Secretary of State',
                          'Foreign Minister', 'Foregin Minster', 'Minister of Foreign Trade and Foreign Investment', 'Foreign Secretary', 'Chairman and General Secretary of the National League of Democracy', 'family photo',
                          'Minister of Foreign Affairs', 'Minister of External Relations', 'Opposition Leader', 'Premier', 'former Prez', 'former PM', 'Sate Dinner', 'working dinner', 'closing session',
                          'Defense Minister', 'Minister for Foreign Affairs', 'Minister of Foreign Affairs', 'Minister of Finance', 'Minister of Foreign Trade', 'Vice-Premeir',
                          'Minister of Environment', 'Minister of Natural Resources', 'Secretary of Foreign Relations', 'State Councilor',
                          'Minister of Power', 'Minister of State for Foreign Affairs','Vice Prez', 'Vice Premier']
    total_meeting_list = ['place holder']
    
    #This tests countries by country name
    for meeting in eval(list_setter):
        for leader in leader_list:
            for country in country_name_list:
                if str(leader) in str(meeting):
                    if str(country) in str(meeting):
                        if str(country) != 'United States':
                            if str(country) != 'American':
                                specific_day = re.compile(r'\w{3,9} \d{1,2}, \d\d\d\d')
                                sd_all_matches = specific_day.findall(meeting)
                                sd = sd_all_matches[0]
                                yao = sd + '\n' + country + '\t' + leader + '\n'
                                searcher = re.compile(r'{}\W+(?:\w+\W+){}?{}'.format(str(leader), "{0,9}", str(country), re.IGNORECASE))
                                backwards_searcher = re.compile(r'(?!.*,){}\W+(?:\w+\W+){}?{}(?!.*,)'.format(str(country), "{0,5}", str(leader), re.IGNORECASE))
                                exact_match = searcher.findall(meeting)
                                other_exact_match = backwards_searcher.findall(meeting)
                                for matches in exact_match:
                                    if ',' not in str(matches):
                                        if 'and ' not in str(matches):
                                            if yao not in total_meeting_list:
                                                if str(country) in yao:
                                                    print(matches)
                                                    print(yao, '1111111')
                                                    total_meeting_list.append(yao)
                                for matches in other_exact_match:
                                    if ',' not in str(matches):
                                        if 'and ' not in str(matches):
                                            if yao not in total_meeting_list:
                                                if str(country) in yao:
                                                    print(matches)
                                                    print(yao, '222222222')
                                                    total_meeting_list.append(yao)
    for diff_country in newest_country_and_list_thats_useable:
        print(new_country_leader_list[diff_country])
    #This tests countries by th names of the leaders
    for diff_country in newest_country_and_list_thats_useable:
        print(new_country_leader_list[diff_country])
        leader_name = str(new_country_leader_list[diff_country][2])
        leader_country = str(new_country_leader_list[diff_country][0])
        other_leader = str(new_country_leader_list[diff_country][-1])
        other_leader_title = str(new_country_leader_list[diff_country][-2])
        leader_title = str(new_country_leader_list[diff_country][1])
        specific_day = re.compile(r'\w{3,9} \d{1,2}, \d\d\d\d')
        if 'President' == str(leader_name):
            pass
        elif 'Prime Minister' == str(leader_name):
            pass
        elif 'President' == str(other_leader):
            pass
        elif 'Prime Minister' == str(other_leader):
            pass
        elif leader_name != other_leader:
            for meeting in eval(list_setter):
                sd_all_matches = specific_day.findall(meeting)
                sd = sd_all_matches[0]
                if str(leader_name) in str(meeting):
                    yao = sd + '\n' + diff_country + '\t' + leader_title + '\n'
                    if yao not in total_meeting_list:
                        print(yao, '3333333333')
                        total_meeting_list.append(yao)
                elif str(other_leader) in str(meeting):
                    yao = sd + '\n' + diff_country + '\t' + other_leader_title + '\n'
                    if yao not in total_meeting_list:
                        print(yao, '444444444')
                        total_meeting_list.append(yao)
        else:
            for meeting in eval(list_setter):
                sd_all_matches = specific_day.findall(meeting)
                sd = sd_all_matches[0]
                if str(leader_name) in str(meeting):
                    yao = sd + '\n' + diff_country + '\t' + leader_title + '\n'
                    if yao not in total_meeting_list:
                        print(yao, '55555555')
                        total_meeting_list.append(yao)
    roman_numeral_list = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X']
    asian_country_last_name_first_list = ['China', 'South Korea', 'Taiwan', 'Vietnam', 'North Korea', 'Singapore', 'Japan']                     
    duplicate_check_meetings_list = ['place_holder']
    #This tests for Leaders last names and titles
    for diff_country in newest_country_and_list_thats_useable:
        leader_name = str(new_country_leader_list[diff_country][2])
        leader_name = leader_name + ' '
        leader_country = str(new_country_leader_list[diff_country][0])
        other_leader = str(new_country_leader_list[diff_country][-1])
        other_leader = other_leader + ' '
        other_leader_title = str(new_country_leader_list[diff_country][-2])
        leader_title = str(new_country_leader_list[diff_country][1])
        specific_day = re.compile(r'\w{3,9} \d{1,2}, \d\d\d\d')                   
        leader_last_name = re.sub('\w{0,10} ', '', leader_name, count=1)
        other_leader_last_name = re.sub('\w{0,10} ', '', other_leader, count=1)
        asian_leader_last_name = re.sub(' \w{0,5}-\w{0,5}| \w{0,10}', '', leader_name, count=1)
        asian_leader_last_name = asian_leader_last_name + ' '
        asian_other_leader_last_name = re.sub(' \w{0,5}-\w{0,5}| \w{0,10}', '', other_leader, count=1)
        asian_other_leader_last_name = asian_other_leader_last_name + ' '
        print(new_country_leader_list[diff_country])
        if len(str(other_leader_last_name)) < 2:
            pass
        elif len(str(leader_last_name)) < 2:
            pass
        elif any(numeral in leader_last_name for numeral in roman_numeral_list):
            pass
        elif any(numeral in other_leader_last_name for numeral in roman_numeral_list):
            pass
        else:
            for meeting in eval(list_setter):
                specific_day = re.compile(r'\w{3,9} \d{1,2}, \d\d\d\d')
                sd_all_matches = specific_day.findall(meeting)
                sd = sd_all_matches[0]
                if str(leader_last_name) in str(meeting):
                    searcher = re.compile(r'{}\W+(?:\w+\W+){}?{}'.format(str(leader_title), "{0,5}", str(leader_last_name), re.IGNORECASE))
                    exact_match = searcher.findall(meeting)
                    for match in exact_match:
                        yao = sd + '\n' + diff_country + '\t' + leader_title + '\n'
                        if yao not in total_meeting_list:
                            if any(country in meeting for country in country_name_list):
                                for country in country_name_list:
                                    if country in yao:
                                        if country in meeting:
                                            total_meeting_list.append(yao)
                            else:
                                print(meeting, leader_last_name)
                                print(yao, '666666')
                                total_meeting_list.append(yao)
                elif str(other_leader_last_name) in str(meeting):
                    searcher = re.compile(r'{}\W+(?:\w+\W+){}?{}'.format(str(other_leader_title), "{0,5}", str(other_leader_last_name), re.IGNORECASE))
                    exact_match = searcher.findall(meeting)
                    for match in exact_match:
                        yao = sd + '\n' + diff_country + '\t' + other_leader_title + '\n'
                        if yao not in total_meeting_list:
                            if any(country in meeting for country in country_name_list):
                                for country in country_name_list:
                                    if country in yao:
                                        if country in meeting:
                                            total_meeting_list.append(yao)
                            else:
                                print(yao, '7777777')                            
                                total_meeting_list.append(yao)
                elif str(asian_leader_last_name) in str(meeting):
                    searcher = re.compile(r'{}\W+(?:\w+\W+){}?{}'.format(str(leader_title), "{0,5}", str(asian_leader_last_name), re.IGNORECASE))
                    exact_match = searcher.findall(meeting)
                    for match in exact_match:
                        yao = sd + '\n' + diff_country + '\t' + leader_title + '\n'                       
                        if yao not in total_meeting_list:
                            if any(country in meeting for country in country_name_list):
                                for country in country_name_list:
                                    if country in yao:
                                        if country in meeting:
                                            total_meeting_list.append(yao)
                            else:
                                print(yao, '88888888')
                                total_meeting_list.append(yao)
                elif str(asian_other_leader_last_name) in str(meeting):
                    searcher = re.compile(r'{}\W+(?:\w+\W+){}?{}'.format(str(other_leader_title), "{0,5}", str(asian_other_leader_last_name), re.IGNORECASE))
                    exact_match = searcher.findall(meeting)
                    for match in exact_match:
                        yao = sd + '\n' + diff_country + '\t' + other_leader_title + '\n'
                        if yao not in total_meeting_list:
                            if any(country in meeting for country in country_name_list):
                                for country in country_name_list:
                                    if country in yao:
                                        if country in meeting:
                                            total_meeting_list.append(yao)
                            else:
                                print(yao, '9999999')
                                total_meeting_list.append(yao)
                elif str('State Dinner') in str(meeting):
                    searcher = re.compile(r'{}\W+(?:\w+\W+){}?{}'.format(str('attend'), "{0,80}", str('a toast'), re.IGNORECASE))
                    exact_match = searcher.findall(meeting)
                    for match in exact_match:
                        yao = sd + '\n' + match + '\n'
                        if yao not in total_meeting_list:
                            print(yao, '1010101')
                            total_meeting_list.append(yao)

    additional_filtration_list = []
    multilateral_country_list = ['G-20', 'G20', 'G8', 'G7', 'ASEAN', 'ASEAN-', 'APEC', 'G-8', 'NATO', 'CARICOM', 'Gulf Cooperation Council', 'GCC', 'G-7', 'Summit of the Americas', 'East Asia Summit', 'Strategic and Economic Dialogue']
    alternative_title_list = ['CEO', 'Crown Prince', 'Defense Minister', 'Deputy Crown Prince', 'Deputy PM', 'Finance Minister', 'Pope', 'Secretary of State',
                          'Foreign Minister', 'Foregin Minster', 'Minister of Foreign Trade and Foreign Investment', 'Foreign Secretary', 'Vice-Premeir',
                          'Minister of Foreign Affairs', 'Minister of External Relations', 'Opposition Leader', 'Premier', 'Chairman and General Secretary of the National League of Democracy',
                          'Defense Minister', 'Minister for Foreign Affairs', 'Minister of Foreign Affairs', 'Minister of Finance', 'Minister of Foreign Trade',
                          'Minister of Environment', 'Minister of Natural Resources', 'Secretary of Foreign Relations', 'State Councilor', 'Premier of the State Council'
                          'Minister of Power', 'Minister of State for Foreign Affairs','Vice Prez', 'Vice Premier', 'former Prez', 'former PM']

    #To make sure title matches with country, so canada/UK dosn't have a president
    for diff_country in newest_country_and_list_thats_useable:
        leader_name = str(new_country_leader_list[diff_country][2])
        leader_country = str(new_country_leader_list[diff_country][0])
        other_leader = str(new_country_leader_list[diff_country][-1])
        other_leader_title = str(new_country_leader_list[diff_country][-2])
        leader_title = str(new_country_leader_list[diff_country][1])
        for meetings in total_meeting_list:
            for country in country_name_list:
                if str(country) in meetings:
                    if str('Afghanistan') in str(meetings):
                        if ('CEO') in str(meetings):
                            additional_filtration_list.append(meetings)
                    elif str('Taoiseach') in str(meetings):
                        if ('Ireland') in str(meetings):
                            additional_filtration_list.append(meetings)
                    elif str('dinner') in str(meetings):
                        if any(multi in meetings for multi in multilateral_country_list):
                            if meetings not in additional_filtration_list:
                                additional_filtration_list.append(meetings)                             
                    elif str(other_leader_title) in meetings:
                        if meetings not in additional_filtration_list:
                            print('xxxxxxxx', meetings, country)
                            additional_filtration_list.append(meetings)
                    elif str(leader_title) in meetings:
                        if meetings not in additional_filtration_list:
                            print('yyyyyyyy', meetings, country, leader_title)
                            additional_filtration_list.append(meetings)
                    else:
                        for title in alternative_title_list:
                            if str(title) in str(meetings):
                                if meetings not in additional_filtration_list:
                                    print('zzzzzzzzzz', meetings, country)
                                    additional_filtration_list.append(meetings)


                        
    #This filters for specific errors
    for meeting in additional_filtration_list:
        meeting = str(meeting)
        prvious_element = duplicate_check_meetings_list[-1]
        last_appended_item = prvious_element
        if meeting[:18] != prvious_element[:18]:
            if 'Mali' in str(meeting) and 'Iraq' in str(prvious_element):
                pass
            elif str('Nigeria') in str(meeting) and str('Niger') in last_appended_item:
                duplicate_check_meetings_list.pop()
                duplicate_check_meetings_list.append(meeting)
            elif str('of the Dominica ') in str(meeting):
                pass
            elif str('Irish') in str(meeting):
                new_ireland = meeting.replace("Irish", "Ireland")
                duplicate_check_meetings_list.append(new_ireland)
            elif str('Dominican Republic') in str(meeting) and str('Dominica') in last_appended_item:
                duplicate_check_meetings_list.pop()
                duplicate_check_meetings_list.append(meeting)
            elif str('China, Peoples Republic of') in str(meeting):
                new_china = meeting.replace("China, Peoples Republic of", "China")
                duplicate_check_meetings_list.append(new_china)
            elif str('Republic of Korea') in str(meeting):
                new_korea = meeting.replace("Republic of Korea", "South Korea")
                duplicate_check_meetings_list.append(new_korea)
            elif str('G8') in str(meeting):
                g8 = meeting.replace("G8", "G-8")
                duplicate_check_meetings_list.append(g8)
            elif str('United States') in str(meeting):
                pass
            elif str('China, Republic of (Taiwan)') in str(meeting):
                new_taiwan = meeting.replace("China, Republic of (Taiwan)", "Taiwan")
                duplicate_check_meetings_list.append(new_taiwan)
            elif str(year) not in str(meeting):
                pass
            elif str('President of the Council of Ministers') in str(meeting):
                pass
            elif str('President of the Government') in str(meeting):
                pass
            else:
                if meeting not in duplicate_check_meetings_list:
                    duplicate_check_meetings_list.append(meeting)
        else:
            for country in country_name_list:
                if str(country) in str(meeting) and str(prvious_element):
                    for title in leader_list:
                        if str(title) in meeting:
                            if str(title) not in str(prvious_element):
                                if meeting not in duplicate_check_meetings_list:
                                    duplicate_check_meetings_list.append(meeting)
                else:
                    if meeting not in duplicate_check_meetings_list:
                        duplicate_check_meetings_list.append(meeting)

    #Further filtration
    new_duplicate_sorting = ['place holder']                         
    for dup in sorted(duplicate_check_meetings_list):
        if dup != new_duplicate_sorting[-1]:
            if str('Federal') in str(dup) and str('Chancellor') in str(new_duplicate_sorting[-1]):
                pass
            elif str('Chancellor') in str(dup):
                if str('Germany') in str(dup):
                    if dup not in new_duplicate_sorting:
                        new_duplicate_sorting.append(dup)                   
                elif str('Merkel') in str(dup):
                    if dup not in new_duplicate_sorting:
                        new_duplicate_sorting.append(dup)
            elif 'Sudan' in str(dup) and 'South Sudan' in str(new_duplicate_sorting[-1]):
                pass
            elif 'Nigeria' in str(dup) and str('Niger') in str(new_duplicate_sorting[-1]):
                new_duplicate_sorting.pop()
                new_duplicate_sorting.append(dup)
            elif 'a href=' in str(dup):
                pass
            elif 'Taoiseach' in str(dup) and str('Ireland	Prime Minister') in str(new_duplicate_sorting[-1]):
                pass
            elif str('CEO') in str(dup):
                if str('Afghanistan') in str(dup):
                    new_duplicate_sorting.append(dup)
                else:
                    pass
            elif 'ession' in str(dup):
                for country in multilateral_country_list:
                    if country not in str(dup):
                        pass
                    else:
                        new_duplicate_sorting.append(dup)
            elif 'eeting' in str(dup):
                for country in multilateral_country_list:
                    if country not in str(dup):
                        pass
                    else:
                        new_duplicate_sorting.append(dup)           
            else:
                if dup not in new_duplicate_sorting:
                    new_duplicate_sorting.append(dup)

    different_country_list_names = ['Niger', 'Ivory Coast']
    new_list = []
    multilateral_event_list = ['photo', 'inner', 'eeting', 'ession']

    #Final Filtration to check if country, demonym or multlateral organization is in string
    for a in new_duplicate_sorting:
        for country in country_dem_list:
            if str(country[1]) in str(a):
                print(a, country[1], 'aaaaaaa')
                for leader in leader_list:
                    if str(leader) in str(a):
                        specific_day = re.compile(r'\w{3,9} \d{1,2}, \d\d\d\d')
                        sd_all_matches = specific_day.findall(a)
                        sd = sd_all_matches[0]
                        all_in_one = sd + '\n' + str(country[0]) + '\t' + str(leader) + '\n'
                        if all_in_one not in new_list:
                            if any(str(event) in str(all_in_one) for event in multilateral_event_list):
                                pass
                            else:
                                new_list.append(all_in_one)
            elif str(country[0]) in str(a):
                if str(a) not in new_list:
                    if any(str(event) in str(a) for event in multilateral_event_list):
                        pass
                    else:
                        new_list.append(a)
            else:
                for country in multilateral_country_list:
                    if str(country) in str(a):
                        if 'President' in str(a):
                            a = a.replace('President', 'Summit Meeting')
                            if a not in new_list:
                                new_list.append(a)
                        elif 'Prime Minister' in str(a):
                            a = a.replace('Prime Minister', 'Summit Meeting')
                            if a not in new_list:
                                new_list.append(a)                                
                        elif str(country) in str(new_list[-1]):
                            specific_day = re.compile(r'\w{3,9} \d{1,2}, \d\d\d\d')
                            sd_all_matches = specific_day.findall(a)
                            sd = sd_all_matches[0]
                            if sd in new_list[-1] and a:
                                pass
                            else:
                                if a not in new_list:
                                    new_list.append(a)
                        else:
                            if a not in new_list:
                                new_list.append(a)
                for diff_count in different_country_list_names:
                    if str(diff_count) in str(a):
                        if a not in new_list:
                            new_list.append(a)                        
                        
    for b in sorted(new_list):
        print(b)
    
    print(len(new_duplicate_sorting))
    print(len(new_list))

    return {'new_list':new_list}

def excel_comparison(country_name_list, new_list):

    QP_wb = openpyxl.load_workbook(excel_file_location)
    Bureaucratic_Exchange_Sheet = QP_wb.get_sheet_by_name('Bureaucratic_Exchange')

    Two_pt_leader_list = ['President', 'Prime Minister', 'Prime Minster', 'King', 'Chancellor', 'Taoiseach', 'Queen', 'Amir', 'Emperor', 'Pope']
    One_pt_leader_list = ['CEO', 'Crown Prince', 'Defense Minister', 'Deputy Crown Prince', 'Deputy PM', 'Finance Minister', 'Secretary of State',
                          'Foreign Minister', 'Foregin Minster', 'Minister of Foreign Trade and Foreign Investment', 'Foreign Secretary',
                          'Minister of Foreign Affairs', 'Minister of External Relations', 'Opposition Leader', 'Premier', 'Chairman and General Secretary of the National League of Democracy',
                          'Defense Minister', 'Minister for Foreign Affairs', 'Minister of Foreign Affairs', 'Minister of Finance', 'Minister of Foreign Trade',
                          'Minister of Environment', 'Minister of Natural Resources', 'Secretary of Foreign Relations', 'State Councilor', 'Premier of the State Council'
                          'Minister of Power', 'Minister of State for Foreign Affairs','Vice Prez', 'Vice Premier', 'Vice-Premeir']
    half_pt_leader_list = ['former Prez', 'former PM', 'President-elect']

    proscessed_list = []
    new_super_filetered_list = new_list

    #This matches the country names from my spreadsheet with the country names returned in this program
    for i in range(3, 201):
        Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
        Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value = 0
        for country in country_name_list:
            if str(Bes.value) == str(country):
                print(str(country), 'aaa')
                for q in new_super_filetered_list:
                    if str(country) in str(q):
                        for leader in Two_pt_leader_list:
                            if str(leader) in str(q):
                                print(str(q), 'ccccc')
                                if len(proscessed_list) < 1:
                                    print(str(q), 'ddddd')
                                    Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                                    proscessed_list.append(q)
                                else:
                                    if str(q) == str(proscessed_list[-1]):
                                        pass
                                    else:
                                        print(str(q), 'eeeee')
                                        Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                                        proscessed_list.append(q)
                        for leader in One_pt_leader_list:
                            if len(proscessed_list) < 1:
                                print(str(q), 'fffff')
                                Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 1
                                proscessed_list.append(q)
                            else:
                                if str(leader) in str(q):
                                    if str(q) == str(proscessed_list[-1]):
                                        pass
                                    else:
                                        print(str(q), 'gggggg')
                                        Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 1
                                        proscessed_list.append(q)
                        for leader in half_pt_leader_list:
                            if len(proscessed_list) < 1:
                                print(str(q), 'hhhhhhhhh')
                                Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += .5
                                proscessed_list.append(q)
                            else:
                                if str(leader) in str(q):
                                    if str(q) != str(proscessed_list[-1]):
                                        print(str(q), 'iiiiii')
                                        Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += .5
                                        proscessed_list.append(q)                                        

            else:
                pass

    caricom = ['Antigua and Barbuda', 'Bahamas, the', 'Barbados', 'Belize', 'Dominica', 'Grenada', 'Guyana', 'Haiti', 'Jamaica', 'Saint Kitts and Nevis', 'Saint Lucia', 'Saint Vincent and the Grenadines', 'Suriname', 'Trinidad and Tobago']
    apec = ['Australia', 'Brunei', 'Canada', 'Chile', 'China', 'Taiwan', 'Indonesia', 'Japan', 'South Korea', 'Malaysia', 'Mexico', 'New Zealand', 'Peru', 'Philippines', 'Papua New Guinea', 'Russia', 'Singapore', 'Thailand', 'Vietnam']
    Summit_of_the_Americas = ['Antigua and Barbuda', 'Argentina', 'Barbados', 'Brazil', 'Belize', 'Bahamas', 'Bolivia', 'Canada', 'Chile', 'Colombia', 'Costa Rica', 'Dominica', 'Dominican Republic', 'Ecuador', 'El Salvador', 'Grenada', 'Guatemala', 'Guyana', 'Haiti', 'Honduras', 'Jamaica', 'Mexico', 'Nicaragua', 'Panama', 'Paraguay', 'Peru', 'Saint Lucia', 'Saint Vincent and the Grenadines', 'Saint Kitts and Nevis', 'Suriname', 'Trinidad and Tobago', 'Uruguay', 'Venezuela']
    G_8_List = ['Russia', 'United Kingdom', 'France', 'Italy', 'Canada', 'Germany', 'Japan']
    G_7_List = ['United Kingdom', 'France', 'Italy', 'Canada', 'Germany', 'Japan']
    G_20_List = ['Argentina', 'Australia', 'Brazil', 'Canada', 'China', 'France', 'Germany', 'India', 'Indonesia', 'Italy', 'Japan', 'South Korea', 'Mexico', 'Russia', 'Saudi Arabia', 'South Africa', 'Turkey', 'United Kingdom']
    NATO = ['Albania', 'Belgium', 'Bulgaria', 'Canada', 'Croatia', 'Czech Republic', 'Denmark', 'Estonia', 'France', 'Germany', 'Greece', 'Hungary', 'Iceland', 'Italy', 'Latvia', 'Lithuania', 'Luxembourg', 'Netherlands', 'Norway', 'Poland', 'Portugal', 'Romania', 'Slovakia', 'Slovenia', 'Spain', 'Turkey', 'United Kingdom']
    GCC = ['Bahrain', 'Kuwait', 'Oman', 'Qatar', 'Saudi Arabia', 'United Arab Emirates']
    East_asia_summit = ['Australia', 'Brunei', 'Cambodia', 'China', 'India', 'Indonesia', 'Japan', 'Laos', 'Malaysia', 'Myanmar', 'New Zealand', 'Philippines', 'Russia', 'Singapore', 'South Korea', 'Thailand', 'Vietnam']
    ASEAN = ['Brunei', 'Cambodia', 'Indonesia', 'Laos', 'Malaysia', 'Myanmar', 'Philippines', 'Singapore', 'Thailand', 'Vietnam', 'Papua New-Guinea', 'East Timor']

    processed_list = []

    for q in new_super_filetered_list:
        if str('State Dinner') in str(q):
            if q in processed_list:
                pass
            else:
                for i in range(3, 201):
                    Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                    if 'Singh' in str(q):
                        if str(Bes.value) == str('India'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
                    if 'Calderon' in str(q):
                        if str(Bes.value) == str('Mexico'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
                    if 'Hu' in str(q):
                        if str(Bes.value) == str('China'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
                    if 'Merkel' in str(q):
                        if str(Bes.value) == str('Germany'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
                    if 'Cameron' in str(q):
                        if str(Bes.value) == str('United Kingdom'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
                    if 'President Lee' in str(q):
                        if str(Bes.value) == str('South Korea'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
                    if 'Hollande' in str(q):
                        if str(Bes.value) == str('France'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
                    if 'Abe' in str(q):
                        if str(Bes.value) == str('Japan'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
                    if 'Xi' in str(q):
                        if str(Bes.value) == str('China'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
                    if 'Trudeau' in str(q):
                        if str(Bes.value) == str('Canada'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
                    if 'Rassmussen' in str(q):
                        if str(Bes.value) == str('Denmark'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
                    if 'Niinistö' in str(q):
                        if str(Bes.value) == str('Finland'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
                    if 'Jóhannsson' in str(q):
                        if str(Bes.value) == str('Iceland'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
                    if 'Solberg' in str(q):
                        if str(Bes.value) == str('Norway'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
                    if 'Löfven' in str(q):
                        if str(Bes.value) == str('Sweden'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
                    if 'Prime Minister Lee' in str(q):
                        if str(Bes.value) == str('Singapore'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
                    if 'Renzi' in str(q):
                        if str(Bes.value) == str('Italy'):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 5
                            processed_list.append(q)
        if str('Strategic and Economic Dialogue') in str(q):
            if q in processed_list:
                pass
            else:
                for i in range(3, 201):
                    Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                    if str(Bes.value) == str('China'):
                        Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                        processed_list.append(q)
        if str('G-20') in str(q):
            if q in processed_list:
                pass
            else:
                for country in G_20_List:
                    for i in range(3, 201):
                        Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                        if str(Bes.value) == str(country):
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
        if str('CARICOM') in str(q):
            if q in processed_list:
                pass
            else:
                for country in caricom:
                    for i in range(3, 201):
                        Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                        if str(Bes.value) == str(country):
                            print(q)
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
        if str('Summit of the Americas') in str(q):
            if q in processed_list:
                pass
            else:
                for country in Summit_of_the_Americas:
                    for i in range(3, 201):
                        Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                        if str(Bes.value) == str(country):
                            print(q)
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
        if str('NATO') in str(q):
            if q in processed_list:
                pass
            else:
                for country in NATO:
                    for i in range(3, 201):
                        Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                        if str(Bes.value) == str(country):
                            print(q)
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
        if str('Gulf Cooperation Council') in str(q):
            if q in processed_list:
                pass
            else:
                for country in GCC:
                    for i in range(3, 201):
                        Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                        if str(Bes.value) == str(country):
                            print(q)
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
        if str('Her Majesty the Queen') in str(q):
            if q in processed_list:
                pass
            else:
                for i in range(3, 201):
                    Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                    if str(Bes.value) == str('United Kingdom'):
                        Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                        processed_list.append(q)
        if str('G-8') in str(q):
            if q in processed_list:
                pass
            elif 'Prime Minister' in str(q):
                pass
            else:
                for country in G_8_List:
                    for i in range(3, 201):
                        Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                        if str(Bes.value) == str(country):
                            print(q)
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
        if str('ASEAN') in str(q):
            if q in processed_list:
                pass
            else:
                for country in ASEAN:
                    for i in range(3, 201):
                        Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                        if str(Bes.value) == str(country):
                            print(q)
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
        if str('East Asia Summit') in str(q): 
            if q in processed_list:
                pass
            else:
                for country in East_asia_summit:
                    for i in range(3, 201):
                        Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                        if str(Bes.value) == str(country):
                            print(q)
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
        if str('G-7') in str(q):
            if q in processed_list:
                pass
            else:
                for country in G_7_List:
                    for i in range(3, 201):
                        Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                        if str(Bes.value) == str(country):
                            print(q)
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)
        if str('APEC') in str(q):
            if q in processed_list:
                pass
            else:
                for country in apec:
                    for i in range(3, 201):
                        Bes = Bureaucratic_Exchange_Sheet.cell(row=i, column=1)
                        if str(Bes.value) == str(country):
                            print(q)
                            Bureaucratic_Exchange_Sheet['{}{}'.format(str(col), str(i))].value += 2
                            processed_list.append(q)

    QP_wb.save(excel_file_location)

    for p in processed_list:
        print(p)
    print("There are", len(proscessed_list), "in this list")
    
def main():
    country_name_list = country_name_acquistion()['country_name_list']
    kerry_daily_meetings_list = daily_sched_search()['kerry_daily_meetings_list']       
    country_dem_list = country_name_acquistion()['country_dem_list']
    
    meeting_search(kerry_daily_meetings_list, country_name_list)
    new_joint_meetings_list = meeting_search(kerry_daily_meetings_list, country_name_list)['new_joint_meetings_list']
    new_PRES_MEETING_LIST = meeting_search(kerry_daily_meetings_list, country_name_list)['new_PRES_MEETING_LIST']
    new_VP_MEETING_LIST = meeting_search(kerry_daily_meetings_list, country_name_list)['new_VP_MEETING_LIST']
    new_list = country_and_leader_getter(new_joint_meetings_list, new_PRES_MEETING_LIST, new_VP_MEETING_LIST, country_name_list, country_dem_list)['new_list']
    
    excel_comparison(country_name_list, new_list)
    
    time2 = time.time()
    print("Total Time to run", time2-time1, 'seconds')

main()

